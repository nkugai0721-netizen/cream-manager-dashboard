"""
Dropbox Excel収支報告 → monthly_pl_history.json 変換スクリプト

使い方:
  python scripts/excel_to_json.py

出力:
  data/monthly_pl_history.json （ダッシュボードが参照）
"""

import openpyxl
import os
import re
import json
import sys
import io
from datetime import datetime

# Windows stdout UTF-8対応
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

BASE_DIR = r"C:\Users\nkuga\Dropbox\CREAM OF THE CROP\00_本部\03_経理\収支報告"

# 店舗ディレクトリ設定
STORE_CONFIG = {
    "せんべろ本店": {
        "dir": "せんべろ風土",
        "pl_pattern": r"せんべろ本店収支報告_(\d{4})\.(\d{1,2})月\.xlsx",
        "mgmt_pattern": r"【せんべろ本店】経営管理_(\d{4})\.(\d{1,2})\.xlsx",
        "format": "senbero",
    },
    "SAN DRIP GARAGE": {
        "dir": "SDG",
        "pl_pattern": r"SDG収支報告_(\d{4})\.(\d{1,2})月\.xlsx",
        "mgmt_pattern": r"【SAN DRIP GARAGE】経営管理_(\d{4})\.(\d{1,2})\.xlsx",
        "format": "sdg_leje",
    },
    "Leje": {
        "dir": "Leje",
        "pl_pattern": r"Leje収支報告_(\d{4})\.(\d{1,2})月\.xlsx",
        "mgmt_pattern": r"【Leje】経営管理_(\d{4})\.(\d{1,2})\.xlsx",
        "format": "sdg_leje",
    },
    "ちゃっきー": {
        "dir": "スナック",
        "pl_pattern": r"(\d{4})\.(\d{1,2})_?(?:スナック|ちゃっきー)収支報告\.xlsx|(?:ちゃっきー|スナック)収支報告_(\d{4})\.(\d{1,2})月\.xlsx",
        "mgmt_pattern": None,
        "format": "snack",
    },
}


def safe_num(val):
    """セル値を数値に変換"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        # ¥マーク、カンマ、スペース除去
        cleaned = val.replace("¥", "").replace(",", "").replace(" ", "").replace("　", "")
        if cleaned == "" or cleaned == "-" or cleaned == "#DIV/0!":
            return 0
        try:
            return float(cleaned)
        except ValueError:
            return 0
    return 0


def cell_val(ws, ref):
    """セル参照から値を取得"""
    return safe_num(ws[ref].value)


def parse_format1_senbero(ws):
    """FORMAT 1: せんべろ本店 収支報告"""
    return {
        "sales": cell_val(ws, "D5"),
        "food": cell_val(ws, "D7"),
        "drink": cell_val(ws, "D8"),
        "other": 0,
        "guests": cell_val(ws, "D11"),
        "avgPrice": cell_val(ws, "D13"),
        "foodCost": cell_val(ws, "D16"),
        "drinkCost": cell_val(ws, "D17"),
        "supplyCost": cell_val(ws, "D19"),
        "varTotal": cell_val(ws, "D18"),
        "labTotal": cell_val(ws, "D24"),
        "labEmployee": cell_val(ws, "D26"),
        "labParttime": cell_val(ws, "D27"),
        "fixedTotal": cell_val(ws, "D28"),
        "opProfit": cell_val(ws, "G30"),
        "isSnack": False,
    }


def parse_format1_sdg_leje(ws):
    """FORMAT 1: SDG / Leje 収支報告"""
    return {
        "sales": cell_val(ws, "D5"),
        "food": cell_val(ws, "D7"),
        "drink": cell_val(ws, "D8"),
        "other": cell_val(ws, "D9") + cell_val(ws, "D10"),  # TO + レンタル
        "guests": cell_val(ws, "D12"),
        "avgPrice": cell_val(ws, "D13"),
        "foodCost": cell_val(ws, "D16"),
        "drinkCost": cell_val(ws, "D17"),
        "supplyCost": cell_val(ws, "D19"),
        "varTotal": cell_val(ws, "D18"),
        "labTotal": cell_val(ws, "D24"),
        "labEmployee": cell_val(ws, "D26"),
        "labParttime": cell_val(ws, "D27"),
        "fixedTotal": cell_val(ws, "D28"),
        "opProfit": cell_val(ws, "G30"),
        "isSnack": False,
    }


def parse_format3_management(ws):
    """FORMAT 3: 経営管理（全店舗共通）"""
    return {
        "sales": cell_val(ws, "B4"),
        "food": cell_val(ws, "B5"),
        "drink": cell_val(ws, "B6"),
        "other": cell_val(ws, "B7"),
        "guests": cell_val(ws, "B9"),
        "avgPrice": cell_val(ws, "B10"),
        "foodCost": cell_val(ws, "B15"),
        "drinkCost": cell_val(ws, "B16"),
        "supplyCost": cell_val(ws, "B17"),
        "varTotal": cell_val(ws, "B18"),
        "labTotal": cell_val(ws, "B23"),
        "labEmployee": cell_val(ws, "B21"),
        "labParttime": cell_val(ws, "B22"),
        "fixedTotal": cell_val(ws, "B41"),
        "opProfit": cell_val(ws, "B44"),
        "isSnack": False,
    }


def parse_snack_old(ws):
    """FORMAT 2A: スナック旧フォーマット"""
    return {
        "sales": cell_val(ws, "D5"),
        "food": 0,
        "drink": 0,
        "other": 0,
        "guests": 0,
        "avgPrice": cell_val(ws, "D13"),
        "foodCost": cell_val(ws, "D16"),
        "drinkCost": cell_val(ws, "D17"),
        "supplyCost": cell_val(ws, "D19"),
        "varTotal": cell_val(ws, "D18"),
        "labTotal": cell_val(ws, "D20"),
        "labEmployee": 0,
        "labParttime": 0,
        "fixedTotal": cell_val(ws, "L17"),
        "opProfit": cell_val(ws, "P8"),
        "isSnack": True,
    }


def parse_snack_new(ws):
    """FORMAT 2B: スナック新フォーマット（月次サマリー）"""
    return {
        "sales": cell_val(ws, "B7"),
        "food": 0,
        "drink": 0,
        "other": 0,
        "guests": 0,
        "avgPrice": 0,
        "foodCost": cell_val(ws, "B14"),  # 仕入れ合計
        "drinkCost": 0,
        "supplyCost": 0,
        "varTotal": cell_val(ws, "B14"),
        "labTotal": cell_val(ws, "B13"),  # キャスト給与合計
        "labEmployee": 0,
        "labParttime": 0,
        "fixedTotal": cell_val(ws, "B26"),
        "opProfit": cell_val(ws, "B39"),
        "isSnack": True,
    }


def compute_ratios(data):
    """比率を計算"""
    s = data["sales"]
    if s > 0:
        data["varRatio"] = data["varTotal"] / s
        data["labRatio"] = data["labTotal"] / s
        data["flRatio"] = (data["varTotal"] + data["labTotal"]) / s
    else:
        data["varRatio"] = 0
        data["labRatio"] = 0
        data["flRatio"] = 0
    return data


def extract_year_month_from_filename(filename, store_name):
    """ファイル名から年月を抽出"""
    config = STORE_CONFIG[store_name]

    # 経営管理ファイル
    if config["mgmt_pattern"]:
        m = re.search(config["mgmt_pattern"], filename)
        if m:
            return int(m.group(1)), int(m.group(2)), "management"

    # 収支報告ファイル
    m = re.search(config["pl_pattern"], filename)
    if m:
        return int(m.group(1)), int(m.group(2)), "pl"

    # スナック: 複数キャプチャグループ対応
    if store_name == "ちゃっきー":
        m = re.search(config["pl_pattern"], filename)
        if m:
            # グループ1,2 または グループ3,4 にマッチ
            y = m.group(1) or m.group(3)
            mo = m.group(2) or m.group(4)
            if y and mo:
                return int(y), int(mo), "pl"
        # フォールバック
        m = re.search(r"(\d{4})\.(\d{1,2})", filename)
        if m:
            return int(m.group(1)), int(m.group(2)), "pl"

    return None, None, None


def process_store(store_name):
    """1店舗の全月データを処理"""
    config = STORE_CONFIG[store_name]
    store_dir = os.path.join(BASE_DIR, config["dir"])
    results = {}

    if not os.path.exists(store_dir):
        print(f"  ディレクトリなし: {store_dir}")
        return results

    # 年別ディレクトリを走査
    for year_dir in sorted(os.listdir(store_dir)):
        year_path = os.path.join(store_dir, year_dir)
        if not os.path.isdir(year_path):
            continue

        for filename in os.listdir(year_path):
            if not filename.endswith(".xlsx") or filename.startswith("~"):
                continue

            year, month, file_type = extract_year_month_from_filename(filename, store_name)
            if year is None:
                continue

            month_key = f"{year}-{month:02d}"
            filepath = os.path.join(year_path, filename)

            # 2026-04以降は経営管理を優先、それ以前は収支報告を優先
            MGMT_PRIORITY_FROM = "2026-04"
            use_mgmt_priority = month_key >= MGMT_PRIORITY_FROM
            if month_key in results:
                existing_source = results[month_key].get("_source")
                if use_mgmt_priority:
                    # 経営管理優先: 既に経営管理データがあればPLはスキップ
                    if existing_source == "management" and file_type == "pl":
                        continue
                else:
                    # 収支報告優先: 既にPLデータがあれば経営管理はスキップ
                    if existing_source == "pl" and file_type == "management":
                        continue

            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)

                if file_type == "management":
                    ws = wb["財務情報"]
                    data = parse_format3_management(ws)
                    data["_source"] = "management"
                elif config["format"] == "senbero":
                    ws = wb[wb.sheetnames[0]]
                    data = parse_format1_senbero(ws)
                    data["_source"] = "pl"
                elif config["format"] == "sdg_leje":
                    ws = wb[wb.sheetnames[0]]
                    data = parse_format1_sdg_leje(ws)
                    data["_source"] = "pl"
                elif config["format"] == "snack":
                    if "月次サマリー" in wb.sheetnames:
                        ws = wb["月次サマリー"]
                        data = parse_snack_new(ws)
                    elif "収支報告" in wb.sheetnames:
                        ws = wb["収支報告"]
                        data = parse_snack_old(ws)
                    else:
                        ws = wb[wb.sheetnames[0]]
                        data = parse_snack_old(ws)
                    data["_source"] = "pl"

                wb.close()

                # 売上0のファイルはスキップ
                if data["sales"] == 0:
                    continue

                compute_ratios(data)

                # 優先ソースで上書き
                if month_key not in results:
                    results[month_key] = data
                elif use_mgmt_priority and file_type == "management":
                    results[month_key] = data
                elif not use_mgmt_priority and file_type == "pl":
                    results[month_key] = data

                print(f"  {month_key} ({file_type}): 売上 ¥{data['sales']:,.0f}")

            except Exception as e:
                print(f"  ERROR {filepath}: {e}")

    return results


def build_monthly_json():
    """全店舗を処理してJSON出力"""
    all_data = {}  # { "2026-07": { "stores": {...}, "total": {...} }, ... }

    for store_name in STORE_CONFIG:
        print(f"\n処理中: {store_name}")
        store_results = process_store(store_name)

        for month_key, data in store_results.items():
            if month_key not in all_data:
                all_data[month_key] = {"stores": {}, "total": {}}

            # _sourceフィールドは出力しない
            clean_data = {k: v for k, v in data.items() if not k.startswith("_")}
            all_data[month_key]["stores"][store_name] = clean_data

    # 全店合計を計算
    sum_fields = ["sales", "food", "drink", "other", "guests", "foodCost",
                  "drinkCost", "supplyCost", "varTotal", "labTotal",
                  "labEmployee", "labParttime", "fixedTotal", "opProfit"]

    for month_key, month_data in all_data.items():
        total = {f: 0 for f in sum_fields}
        total["isSnack"] = False

        for store_data in month_data["stores"].values():
            for f in sum_fields:
                total[f] += store_data.get(f, 0)

        # 全店平均客単価
        if total["guests"] > 0:
            total["avgPrice"] = round(total["sales"] / total["guests"])
        else:
            total["avgPrice"] = 0

        compute_ratios(total)
        month_data["total"] = total

    # 月でソートして出力
    sorted_data = dict(sorted(all_data.items()))

    # 出力先
    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "monthly_pl_history.json")

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(sorted_data, f, ensure_ascii=False, indent=2)

    print(f"\n出力: {out_path}")
    print(f"月数: {len(sorted_data)}")
    print(f"期間: {min(sorted_data.keys())} ～ {max(sorted_data.keys())}")

    # サマリー
    for month_key in sorted(sorted_data.keys()):
        stores = list(sorted_data[month_key]["stores"].keys())
        total_sales = sorted_data[month_key]["total"]["sales"]
        print(f"  {month_key}: {', '.join(stores)} → 合計 ¥{total_sales:,.0f}")


if __name__ == "__main__":
    build_monthly_json()
